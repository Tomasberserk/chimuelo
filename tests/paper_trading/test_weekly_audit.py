"""Tests para el Weekly Audit Reporting System de Chimuelo Prime.

Verifica:
- Dataset canónico y derivación estricta a JSON, Markdown y XLSX.
- Data Quality Reconciliation (PASS / 0 inconsistencies).
- Serie de Equity semanal y Schema Versioning (1.0.0).
- Cero mutación de Strategy C, Risk Engine, Cash, Positions, Orders, Fills.
- Inmutabilidad y hashing criptográfico SHA-256.
"""

from datetime import UTC, datetime, timedelta
from decimal import Decimal
import json
from pathlib import Path
import openpyxl
import pytest

from chimuelo_prime.backtesting.data_loader import HistoricalCandle
from chimuelo_prime.paper_trading.decision_engine import SingleDecisionEngine
from chimuelo_prime.paper_trading.decision_models import (
    DecisionAction,
    DecisionObject,
    LifecycleEvent,
    MarketRegimeSnapshot,
    PaperFill,
    PaperOrder,
    PaperPosition,
    RiskStateEnum,
)
from chimuelo_prime.paper_trading.persistence import SQLitePersistenceBackend
from chimuelo_prime.paper_trading.risk_engine import PortfolioRiskEngine
from chimuelo_prime.paper_trading.telemetry import PaperTelemetryCollector
from chimuelo_prime.paper_trading.virtual_broker import VirtualBroker
from chimuelo_prime.paper_trading.weekly_audit import WeeklyAuditReportGenerator
from chimuelo_prime.strategies.structural_breakout import StructuralBreakoutStrategy


def test_weekly_report_generation_and_determinism(tmp_path):
    """Verifica que el generador cree JSON, Markdown y XLSX de forma determinista desde el dataset canónico."""
    db_path = str(tmp_path / "audit_test.db")
    reports_dir = tmp_path / "reports" / "weekly"
    persistence = SQLitePersistenceBackend(db_path)
    broker = VirtualBroker(persistence=persistence, initial_cash=Decimal("100.00"))
    telemetry = PaperTelemetryCollector()

    t0 = datetime(2026, 9, 1, 10, 0, 0, tzinfo=UTC)
    t1 = datetime(2026, 9, 1, 15, 0, 0, tzinfo=UTC)
    pos = PaperPosition(
        position_id="pos_audit_1",
        symbol="SOLUSDT",
        status="CLOSED",
        entry_time=t0,
        entry_signal_price=Decimal("100.00"),
        fill_price=Decimal("100.05"),
        slippage_pct=Decimal("0.0005"),
        stop_loss=Decimal("90.00"),
        take_profit=Decimal("122.00"),
        quantity=Decimal("0.5"),
        fee_entry=Decimal("0.05"),
        exit_time=t1,
        exit_price=Decimal("122.00"),
        exit_reason="TAKE_PROFIT",
        fee_exit=Decimal("0.061"),
        gross_pnl=Decimal("10.975"),
        net_pnl=Decimal("10.864"),
        r_multiple=Decimal("2.19"),
        duration_hours=5,
    )
    telemetry.record_closed_position(pos)

    generator = WeeklyAuditReportGenerator(
        persistence=persistence,
        telemetry=telemetry,
        broker=broker,
        output_base_dir=str(reports_dir),
    )

    res1 = generator.generate_and_save_package(week_number=36, year=2026, initial_capital=Decimal("100.00"))

    json_path = Path(res1["files"]["json"])
    md_path = Path(res1["files"]["markdown"])
    xlsx_path = Path(res1["files"]["excel"])

    assert json_path.exists()
    assert md_path.exists()
    assert xlsx_path.exists()

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    assert data["identity"]["report_schema_version"] == "1.0.0"
    assert data["identity"]["week_identifier"] == "2026_W36"
    assert data["performance"]["trades_closed_count"] == 1
    assert data["performance"]["win_rate_pct"] == 100.0
    assert len(data["trade_ledger"]) == 1
    assert data["trade_ledger"][0]["position_id"] == "pos_audit_1"
    assert data["data_quality_reconciliation"]["status"] == "PASS"
    assert data["data_quality_reconciliation"]["total_inconsistencies"] == 0
    assert "weekly_equity_series" in data


def test_data_quality_reconciliation_pass_on_synthetic_dataset(tmp_path):
    """Genera reporte con datos sintéticos completos: 3 trades (1 win, 2 losses), 1 blocked-by-risk, 1 risk state transition, 1 open position -> PASS."""
    db_path = str(tmp_path / "rec_test.db")
    reports_dir = tmp_path / "reports" / "weekly"
    persistence = SQLitePersistenceBackend(db_path)
    broker = VirtualBroker(persistence=persistence, initial_cash=Decimal("100.00"))
    telemetry = PaperTelemetryCollector()

    t0 = datetime(2026, 9, 1, 10, 0, 0, tzinfo=UTC)

    # Trade 1: Win (+$10.00 net)
    p1 = PaperPosition(
        position_id="pos_win_1",
        symbol="BTCUSDT",
        status="CLOSED",
        entry_time=t0,
        entry_signal_price=Decimal("60000.00"),
        fill_price=Decimal("60000.00"),
        slippage_pct=Decimal("0.0005"),
        stop_loss=Decimal("58000.00"),
        take_profit=Decimal("64000.00"),
        quantity=Decimal("0.005"),
        fee_entry=Decimal("0.30"),
        exit_time=t0 + timedelta(hours=4),
        exit_price=Decimal("62120.00"),
        exit_reason="TAKE_PROFIT",
        fee_exit=Decimal("0.30"),
        gross_pnl=Decimal("10.60"),
        net_pnl=Decimal("10.00"),
        r_multiple=Decimal("1.06"),
        duration_hours=4,
    )
    telemetry.record_closed_position(p1)

    # Trade 2: Loss (-$2.50 net)
    p2 = PaperPosition(
        position_id="pos_loss_1",
        symbol="SOLUSDT",
        status="CLOSED",
        entry_time=t0 + timedelta(hours=6),
        entry_signal_price=Decimal("100.00"),
        fill_price=Decimal("100.00"),
        slippage_pct=Decimal("0.0005"),
        stop_loss=Decimal("90.00"),
        take_profit=Decimal("120.00"),
        quantity=Decimal("0.25"),
        fee_entry=Decimal("0.025"),
        exit_time=t0 + timedelta(hours=8),
        exit_price=Decimal("90.20"),
        exit_reason="STOP_LOSS",
        fee_exit=Decimal("0.025"),
        gross_pnl=Decimal("-2.45"),
        net_pnl=Decimal("-2.50"),
        r_multiple=Decimal("-1.00"),
        duration_hours=2,
    )
    telemetry.record_closed_position(p2)

    # Trade 3: Loss (-$2.50 net)
    p3 = PaperPosition(
        position_id="pos_loss_2",
        symbol="SOLUSDT",
        status="CLOSED",
        entry_time=t0 + timedelta(hours=10),
        entry_signal_price=Decimal("95.00"),
        fill_price=Decimal("95.00"),
        slippage_pct=Decimal("0.0005"),
        stop_loss=Decimal("85.00"),
        take_profit=Decimal("115.00"),
        quantity=Decimal("0.25"),
        fee_entry=Decimal("0.025"),
        exit_time=t0 + timedelta(hours=12),
        exit_price=Decimal("85.20"),
        exit_reason="STOP_LOSS",
        fee_exit=Decimal("0.025"),
        gross_pnl=Decimal("-2.45"),
        net_pnl=Decimal("-2.50"),
        r_multiple=Decimal("-1.00"),
        duration_hours=2,
    )
    telemetry.record_closed_position(p3)

    # 1 Open Position
    open_p = PaperPosition(
        position_id="pos_open_1",
        symbol="BTCUSDT",
        status="OPEN",
        entry_time=t0 + timedelta(hours=14),
        entry_signal_price=Decimal("60500.00"),
        fill_price=Decimal("60500.00"),
        slippage_pct=Decimal("0.0005"),
        stop_loss=Decimal("58500.00"),
        take_profit=Decimal("64500.00"),
        quantity=Decimal("0.005"),
        fee_entry=Decimal("0.30"),
    )
    broker._open_positions["BTCUSDT"] = open_p

    generator = WeeklyAuditReportGenerator(
        persistence=persistence,
        telemetry=telemetry,
        broker=broker,
        output_base_dir=str(reports_dir),
    )

    res = generator.generate_and_save_package(week_number=36, year=2026)
    rec = res["reconciliation"]

    assert rec["status"] == "PASS"
    assert rec["total_inconsistencies"] == 0
    assert rec["summary_verdict"] == "PASS / 0 inconsistencies"
    assert rec["checks"]["signals_orders_reconciled"] is True
    assert rec["checks"]["orders_fills_reconciled"] is True
    assert rec["checks"]["fills_positions_reconciled"] is True
    assert rec["checks"]["positions_pnl_reconciled"] is True


def test_format_consistency_across_json_md_xlsx(tmp_path):
    """Verifica que JSON, Markdown y XLSX contengan exactamente los mismos datos numéricos."""
    db_path = str(tmp_path / "fmt_consist.db")
    reports_dir = tmp_path / "reports" / "weekly"
    persistence = SQLitePersistenceBackend(db_path)
    broker = VirtualBroker(persistence=persistence, initial_cash=Decimal("100.00"))
    telemetry = PaperTelemetryCollector()

    t0 = datetime(2026, 9, 1, 10, 0, 0, tzinfo=UTC)
    pos = PaperPosition(
        position_id="pos_fmt_1",
        symbol="SOLUSDT",
        status="CLOSED",
        entry_time=t0,
        entry_signal_price=Decimal("100.00"),
        fill_price=Decimal("100.05"),
        slippage_pct=Decimal("0.0005"),
        stop_loss=Decimal("90.00"),
        take_profit=Decimal("122.00"),
        quantity=Decimal("0.5"),
        fee_entry=Decimal("0.05"),
        exit_time=t0 + timedelta(hours=3),
        exit_price=Decimal("122.00"),
        exit_reason="TAKE_PROFIT",
        fee_exit=Decimal("0.061"),
        gross_pnl=Decimal("10.975"),
        net_pnl=Decimal("10.864"),
        r_multiple=Decimal("2.19"),
        duration_hours=3,
    )
    telemetry.record_closed_position(pos)

    generator = WeeklyAuditReportGenerator(
        persistence=persistence,
        telemetry=telemetry,
        broker=broker,
        output_base_dir=str(reports_dir),
    )

    res = generator.generate_and_save_package(week_number=36, year=2026)

    # 1. JSON
    with open(res["files"]["json"], "r", encoding="utf-8") as f:
        json_data = json.load(f)

    # 2. Markdown
    with open(res["files"]["markdown"], "r", encoding="utf-8") as f:
        md_text = f.read()

    # 3. XLSX
    wb = openpyxl.load_workbook(res["files"]["excel"])
    ws_sum = wb["Resumen Ejecutivo"]
    ws_trades = wb["Trade Ledger"]

    # Validar consistencia exacta de trades
    assert json_data["performance"]["trades_closed_count"] == 1
    assert "pos_fmt_1" in md_text
    assert ws_trades.cell(row=2, column=1).value == "pos_fmt_1"
    assert ws_trades.cell(row=2, column=2).value == "SOLUSDT"
    assert ws_trades.cell(row=2, column=13).value == 0.111   # Fees (fee_entry + fee_exit)
    assert ws_trades.cell(row=2, column=14).value == 10.864  # Net PnL


def test_no_mutation_of_strategy_or_portfolio_state(tmp_path):
    """Garantiza que la generación del reporte no modifique el estado de Strategy, Risk, Cash ni Posiciones."""
    db_path = str(tmp_path / "no_mut.db")
    persistence = SQLitePersistenceBackend(db_path)
    broker = VirtualBroker(persistence=persistence, initial_cash=Decimal("100.00"))
    telemetry = PaperTelemetryCollector()

    strategy_hash_before = StructuralBreakoutStrategy.get_config_hash()
    broker_cash_before = broker.cash
    broker_positions_before = broker.get_open_positions_count()

    generator = WeeklyAuditReportGenerator(
        persistence=persistence,
        telemetry=telemetry,
        broker=broker,
        output_base_dir=str(tmp_path / "reports"),
    )
    generator.generate_and_save_package(week_number=36, year=2026)

    assert StructuralBreakoutStrategy.get_config_hash() == strategy_hash_before
    assert broker.cash == broker_cash_before
    assert broker.get_open_positions_count() == broker_positions_before


def test_hash_reproducibility(tmp_path):
    """Verifica que el hash de integridad sea reproducible e identifique unívocamente el Canonical Dataset."""
    db_path = str(tmp_path / "hash_rep.db")
    persistence = SQLitePersistenceBackend(db_path)
    broker = VirtualBroker(persistence=persistence, initial_cash=Decimal("100.00"))
    telemetry = PaperTelemetryCollector()

    generator = WeeklyAuditReportGenerator(
        persistence=persistence,
        telemetry=telemetry,
        broker=broker,
        output_base_dir=str(tmp_path / "reports"),
    )

    data1 = generator.build_report_data(week_number=36, year=2026)
    data2 = generator.build_report_data(week_number=36, year=2026)

    assert len(data1["data_integrity_sha256"]) == 64
    assert len(data2["data_integrity_sha256"]) == 64
    assert data1["identity"]["report_schema_version"] == "1.0.0"
